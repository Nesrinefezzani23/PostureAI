from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import Session, RawMeasure, PosturalAnalysis, Alerte
from rest_framework import generics, status
from rest_framework.response import Response
from .serializers import RegisterSerializer, RawMeasureSerializer
from rest_framework.permissions import AllowAny

class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    permission_classes = (AllowAny,)
    serializer_class = RegisterSerializer

from .ai_engine import analyze_posture_data
from django.utils import timezone
from datetime import timedelta
from django.db.models import Avg

import csv
from django.http import HttpResponse, JsonResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from .models import Session, RawMeasure, PosturalAnalysis

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

def export_csv(request):
    """Exporte les analyses posturales au format CSV enrichi."""
    analyses = PosturalAnalysis.objects.all().order_by('-timestamp_analyse')
    if not analyses.exists():
        return HttpResponse("Aucune donnée disponible.")
        
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rapport_posture_ai.csv"'
    
    writer = csv.writer(response)
    # Metadata
    user = analyses.first().session.user
    first_date = analyses.last().timestamp_analyse.strftime('%Y-%m-%d')
    last_date = analyses.first().timestamp_analyse.strftime('%Y-%m-%d')
    
    writer.writerow(['Utilisateur', user.username])
    writer.writerow(['Periode', f"{first_date} au {last_date}"])
    writer.writerow([]) # Empty row
    writer.writerow(['Date', 'Score', 'Zone de Tension', 'Statut', 'Recommandation'])
    
    for analyse in analyses:
        writer.writerow([
            analyse.timestamp_analyse.strftime('%Y-%m-%d %H:%M'),
            analyse.score_posture,
            analyse.zone_tension,
            analyse.statut,
            analyse.recommandation
        ])
    return response

def export_excel(request):
    """Génère un rapport Excel (.xlsx) stylisé avec couleurs conditionnelles et auto-ajustement."""
    analyses = PosturalAnalysis.objects.all().order_by('-timestamp_analyse')
    if not analyses.exists():
        return HttpResponse("Aucune donnée disponible.")
        
    username = analyses.first().session.user.username
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport PostureAI"
    
    # Styles
    primary_color = "D4A373"
    header_fill = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
    font_bold = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center")
    
    # Add Logo
    img = ExcelImage('dashboard/static/dashboard/img/Logo.png')
    img.width = 100
    img.height = 100
    ws.add_image(img, 'A1')
    
    # Metadata
    first_date = analyses.last().timestamp_analyse.strftime('%d/%m/%Y')
    last_date = analyses.first().timestamp_analyse.strftime('%d/%m/%Y')
    
    ws['C2'] = "Rapport d'Analyse Posturale"
    ws['C2'].font = Font(size=18, bold=True)
    ws['C3'] = f"Utilisateur : {username}"
    ws['C4'] = f"Période : {first_date} - {last_date}"
    
    # Extra empty row for spacing
    ws.append([])
    
    # Header row
    headers = ['Date', 'Score', 'Zone de Tension', 'Statut', 'Recommandation']
    ws.append(headers)
    for cell in ws[6]:
        cell.fill = header_fill
        cell.font = Font(size=12, bold=True, color="FFFFFF")
        cell.alignment = center_align
    
    # Data rows
    for analyse in analyses:
        # Status colors
        status_color = "70C157" if analyse.statut == 'vert' else ("F4A261" if analyse.statut == 'orange' else "E76F51")
        status_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
        
        row = [
            analyse.timestamp_analyse.strftime('%d/%m/%Y %H:%M'),
            analyse.score_posture,
            analyse.zone_tension,
            analyse.statut.upper(),
            analyse.recommandation
        ]
        ws.append(row)
        ws.cell(row=ws.max_row, column=4).fill = status_fill
        ws.cell(row=ws.max_row, column=4).alignment = center_align
    
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rapport_posture_{username}.xlsx"'
    wb.save(response)
    return response

def export_pdf(request):
    """Génère un rapport médical au format PDF moderne, chaleureux avec pagination."""
    analyses = PosturalAnalysis.objects.all().order_by('-timestamp_analyse')
    if not analyses.exists():
        return HttpResponse("Aucune donnée disponible.")
    
    username = analyses.first().session.user.username
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="rapport_posture_{username}.pdf"'
    
    p = canvas.Canvas(response, pagesize=A4)
    # Warm Palette
    warm_bg = colors.HexColor("#fffcf9")
    header_bg = colors.HexColor("#f8f0e5")
    primary_warm = colors.HexColor("#d4a373")
    text_warm = colors.HexColor("#5b4636")
    
    analyses = PosturalAnalysis.objects.all().order_by('-timestamp_analyse')
    if not analyses.exists():
        p.drawString(100, 700, "Aucune donnée disponible.")
        p.save()
        return response
    
    # Context data
    user = analyses.first().session.user
    first_date = analyses.last().timestamp_analyse.strftime('%d/%m/%Y')
    last_date = analyses.first().timestamp_analyse.strftime('%d/%m/%Y')
    
    page_num = 1
    def draw_header_and_footer():
        # Background
        p.setFillColor(warm_bg)
        p.rect(0, 0, 600, 850, stroke=0, fill=1)
        
        # Header (Le bandeau de fond)
        p.setFillColor(header_bg)
        p.rect(0, 760, 600, 82, stroke=0, fill=1)
        
        # Logo (Placé à Y=750 avec une hauteur de 90)
        logo_path = 'dashboard/static/dashboard/img/Logo.png'
        p.drawImage(logo_path, 35, 750, width=90, height=90, mask='auto')
        
        p.setFillColor(text_warm)
        
        # --- ALIGNEMENT VERTICAL ---
        # Titre : descendu de 810 à 800
        p.setFont("Helvetica-Bold", 22)
        p.drawString(140, 800, "RAPPORT D'ANALYSE POSTURALE")
        
        # Sous-titre : descendu de 785 à 775
        p.setFont("Helvetica", 10)
        p.drawString(140, 775, f"Rapport : {user.username.capitalize()} | {first_date} - {last_date}")
        # ---------------------------
        
        # Footer
        p.setFont("Helvetica-Oblique", 9)
        p.drawCentredString(300, 20, f"Page {page_num} | Généré par PostureAI - Votre bien-être est notre priorité.")

    draw_header_and_footer()
    
    # Reduced gap between header and title
    y = 730 
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, y, "Historique des Analyses")
    
    # Increased gap between title and table
    y -= 45 
    
    # Table Header
    p.setFillColor(primary_warm)
    p.rect(50, y, 500, 25, stroke=0, fill=1)
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(60, y + 8, "DATE & HEURE")
    p.drawString(180, y + 8, "SCORE")
    p.drawString(260, y + 8, "ZONE DE TENSION")
    p.drawString(420, y + 8, "STATUT")
    
    y -= 25
    p.setFont("Helvetica", 10)
    
    for i, analyse in enumerate(analyses[:40]):
        if y < 100:
            p.showPage()
            page_num += 1
            draw_header_and_footer()
            y = 700
        
        if i % 2 == 0:
            p.setFillColor(colors.white)
        else:
            p.setFillColor(colors.HexColor("#faeddd"))
        p.rect(50, y - 5, 500, 20, stroke=0, fill=1)
        
        p.setFillColor(text_warm)
        p.drawString(60, y, analyse.timestamp_analyse.strftime('%d/%m %H:%M'))
        p.drawString(180, y, str(analyse.score_posture))
        p.drawString(260, y, str(analyse.zone_tension))
        
        status_color = colors.HexColor("#606c38") if analyse.statut == 'vert' else (colors.HexColor("#bc6c25") if analyse.statut == 'orange' else colors.HexColor("#bc4749"))
        p.setFillColor(status_color)
        p.drawString(420, y, analyse.statut.upper())
        y -= 20
    
    p.showPage()
    p.save()
    return response

from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import User
from .models import Profile
from rest_framework.decorators import api_view

def landing(request):
    return render(request, 'dashboard/landing.html')

def signin(request):
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        user = authenticate(request, username=u, password=p)
        if user is not None:
            login(request, user)
            return redirect('home')
    return render(request, 'dashboard/login.html')

def signup(request):
    if request.method == 'POST':
        u = request.POST.get('username')
        e = request.POST.get('email')
        p = request.POST.get('password')
        t = request.POST.get('taille')
        w = request.POST.get('poids_kg')
        a = request.POST.get('activite', 'sedentaire')
        patho = request.POST.get('pathologies', '')

        user = User.objects.create_user(username=u, email=e, password=p)
        Profile.objects.create(
            user=user,
            taille=t if t else None,
            poids_kg=w if w else None,
            activite=a,
            pathologies=patho
        )
        login(request, user)
        return redirect('home')
    return render(request, 'dashboard/register.html')

def signout(request):
    logout(request)
    return redirect('signin')

@login_required
def profile_settings(request):
    if request.method == 'POST':
        user = request.user
        user.profile.taille = request.POST.get('taille')
        user.profile.poids_kg = request.POST.get('poids_kg')
        user.profile.activite = request.POST.get('activite')
        user.profile.pathologies = request.POST.get('pathologies')
        user.profile.save()
        return redirect('home')
    
    context = {
        'user': request.user,
        'activite_choices': Profile.ACTIVITE_CHOICES
    }
    return render(request, 'dashboard/profile.html', context)

@api_view(['POST'])
def receive_data(request):
    """
    API pour l'ESP32.
    """
    serializer = RawMeasureSerializer(data=request.data)
    if serializer.is_valid():
        measure = serializer.save() 
        analysis_results = analyze_posture_data(measure)
        
        PosturalAnalysis.objects.create(
            measure=measure,
            session=measure.session,
            score_posture=analysis_results['score'],
            deviation_dos=analysis_results['deviation'],
            deviation_cou=analysis_results['deviation_cou'], 
            symetrie_pression=analysis_results['symetrie_pression'], 
            zone_tension=analysis_results['zone_tension'],
            statut=analysis_results['statut'],
            immobilite_min=0,
            recommandation=analysis_results['recommandation']
        )
        return Response({"status": "success"}, status=status.HTTP_201_CREATED)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

from django.core.paginator import Paginator

@login_required
def home(request):
    """
    Affiche le dashboard avec les dernières données et l'utilisateur de la session.
    """
    # 1. Statistiques globales pour l'utilisateur connecté
    total_sessions = Session.objects.filter(user=request.user).count()
    
    # 2. Dernière mesure brute reçue pour cet utilisateur
    derniere_mesure = RawMeasure.objects.filter(session__user=request.user).order_by('-timestamp').first()

    # 3. Analyse IA la plus récente pour cet utilisateur
    derniere_analyse = PosturalAnalysis.objects.filter(session__user=request.user).order_by('-timestamp_analyse').first()

    # 4. Récupération du nom de l'utilisateur
    nom_utilisateur = request.user.username

    context = {
        'total_sessions': total_sessions,
        'derniere_mesure': derniere_mesure,
        'derniere_analyse': derniere_analyse,
        'nom_utilisateur': nom_utilisateur,
        'status_ia': "Actif",
        'activite_choices': Profile.ACTIVITE_CHOICES
    }
    return render(request, 'dashboard/index.html', context)

@login_required
def historique(request):
    """
    Vue pour l'historique compact avec graphique hebdomadaire.
    """
    aujourdhui = timezone.now().date()
    # Création propre de la liste des 7 derniers jours (du plus vieux au plus récent)
    derniers_7_jours = [aujourdhui - timedelta(days=i) for i in range(6, -1, -1)]
    
    labels_jours = []
    scores_jours = []
    
    # Boucle pour l'utilisateur connecté
    for jour in derniers_7_jours:
        labels_jours.append(jour.strftime('%d %b'))
        
        stat_du_jour = PosturalAnalysis.objects.filter(
            session__user=request.user,
            timestamp_analyse__date=jour
        ).aggregate(moyenne_score=Avg('score_posture'))
        
        score_moyen = stat_du_jour['moyenne_score']
        if score_moyen is not None:
            scores_jours.append(round(score_moyen, 1))
        else:
            scores_jours.append(0) 

    # Pagination des analyses pour l'utilisateur connecté
    analyses_list = PosturalAnalysis.objects.filter(session__user=request.user).order_by('-timestamp_analyse')
    paginator = Paginator(analyses_list, 5)
    page_number = request.GET.get('page')
    dernieres_analyses = paginator.get_page(page_number)

    context = {
        'labels_jours': labels_jours,
        'scores_jours': scores_jours,
        'dernieres_analyses': dernieres_analyses,
    }
    return render(request, 'dashboard/historique.html', context)


@login_required
def get_alertes(request):
    alertes = Alerte.objects.filter(
        user=request.user,
        lue=False
    ).select_related('analysis').order_by('-declenchee_at')[:20]

    data = [{
        'id': str(a.id_uuid),
        'type': a.type_alerte,
        'message': a.message,
        'statut': a.analysis.statut,
        'score': a.analysis.score_posture,
        'zone': a.analysis.zone_tension,
        'heure': a.declenchee_at.strftime('%H:%M:%S'),
    } for a in alertes]

    # Marquer comme lues
    alertes.update(lue=True)

    return JsonResponse({'alertes': data})